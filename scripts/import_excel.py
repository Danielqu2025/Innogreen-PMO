#!/usr/bin/env python3
"""
从 Excel 导入数据到数据库
支持: 工作阶段划分_v2.xlsx (5-sheet)
"""

import sqlite3
import sys
from pathlib import Path
from typing import List, Tuple, Optional

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

def import_from_excel(db_path: str, excel_path: str) -> bool:
    """从 Excel 导入数据"""
    print(f"📖 读取 Excel: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ 无法打开 Excel 文件: {e}")
        return False

    conn = sqlite3.connect(db_path)
    conn.execute('PRAGMA journal_mode=WAL;')
    cursor = conn.cursor()

    imported = {'stages': 0, 'tasks': 0, 'dependencies': 0, 'pitfalls': 0}

    # Sheet1: 阶段定义
    if 'Sheet1' in wb.sheetnames or '阶段定义' in wb.sheetnames:
        sheet_name = '阶段定义' if '阶段定义' in wb.sheetnames else 'Sheet1'
        sheet = wb[sheet_name]
        imported['stages'] = import_stages(cursor, sheet)
        print(f"   ✓ 导入 {imported['stages']} 个阶段")

    # Sheet2: 任务明细
    if 'Sheet2' in wb.sheetnames or '任务明细' in wb.sheetnames:
        sheet_name = '任务明细' if '任务明细' in wb.sheetnames else 'Sheet2'
        sheet = wb[sheet_name]
        imported['tasks'] = import_tasks(cursor, sheet)
        print(f"   ✓ 导入 {imported['tasks']} 个任务")

    # Sheet3: 避坑指南
    if 'Sheet3' in wb.sheetnames or '避坑指南' in wb.sheetnames:
        sheet_name = '避坑指南' if '避坑指南' in wb.sheetnames else 'Sheet3'
        sheet = wb[sheet_name]
        imported['pitfalls'] = import_pitfalls(cursor, sheet)
        print(f"   ✓ 导入 {imported['pitfalls']} 条避坑")

    # Sheet5: 任务依赖 (如果存在)
    if 'Sheet5' in wb.sheetnames or '任务依赖' in wb.sheetnames:
        sheet_name = '任务依赖' if '任务依赖' in wb.sheetnames else 'Sheet5'
        sheet = wb[sheet_name]
        imported['dependencies'] = import_dependencies(cursor, sheet)
        print(f"   ✓ 导入 {imported['dependencies']} 条依赖")

    conn.commit()
    conn.close()

    print(f"\n📊 导入汇总: {imported}")
    return True


def import_stages(cursor: sqlite3.Cursor, sheet) -> int:
    """导入阶段数据"""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0

    # 跳过标题行
    header = [str(h).strip() if h else '' for h in rows[0]]
    print(f"   阶段表头: {header[:8]}")

    count = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue

        # 查找列索引
        stage_id_idx = find_column(header, ['stage_id', '阶段ID', '序号', 'id'])
        stage_name_idx = find_column(header, ['stage_name', '阶段名称', '阶段', '名称'])
        owner_idx = find_column(header, ['primary_owner', '责任方', '主导方'])
        critical_idx = find_column(header, ['critical_path', '关键路径', '关键度'])
        days_idx = find_column(header, ['default_days', '默认工期', '工期'])
        desc_idx = find_column(header, ['description', '描述', '说明'])
        sort_idx = find_column(header, ['sort_order', '排序', '顺序'])

        if stage_name_idx is None:
            continue

        stage_name = str(row[stage_name_idx]).strip() if row[stage_name_idx] else ''
        if not stage_name:
            continue

        stage_id = int(row[stage_id_idx]) if stage_id_idx and row[stage_id_idx] else None
        owner = str(row[owner_idx]).strip() if owner_idx and row[owner_idx] else '客户主导'
        critical = str(row[critical_idx]).strip() if critical_idx and row[critical_idx] else '🟢'
        days = int(row[days_idx]) if days_idx and row[days_idx] else 0
        description = str(row[desc_idx]).strip() if desc_idx and row[desc_idx] else ''
        sort_order = int(row[sort_idx]) if sort_idx and row[sort_idx] else count + 1

        # 映射关键路径符号
        critical = map_critical_path(critical)

        try:
            cursor.execute('''
                INSERT OR IGNORE INTO stage_map
                (stage_name, primary_owner, critical_path, default_days, description, sort_order)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (stage_name, owner, critical, days, description, sort_order))
            if cursor.rowcount > 0:
                count += 1
        except sqlite3.Error as e:
            print(f"   ⚠️ 阶段 '{stage_name}' 导入失败: {e}")

    return count


def import_tasks(cursor: sqlite3.Cursor, sheet) -> int:
    """导入任务数据"""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0

    header = [str(h).strip() if h else '' for h in rows[0]]
    print(f"   任务表头: {header[:10]}")

    count = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue

        stage_id_idx = find_column(header, ['stage_id', '阶段ID', '所属阶段'])
        task_name_idx = find_column(header, ['task_name', '任务名称', '任务'])
        seq_idx = find_column(header, ['seq', '序号', '顺序'])
        days_idx = find_column(header, ['default_days', '默认工期', '工期'])
        critical_idx = find_column(header, ['critical_path', '关键路径', '关键度'])
        owner_idx = find_column(header, ['owner', '责任方', '主导'])
        desc_idx = find_column(header, ['description', '描述', '说明'])
        accept_idx = find_column(header, ['acceptance_criteria', '验收标准', '标准'])
        sort_idx = find_column(header, ['sort_order', '排序'])

        if task_name_idx is None:
            continue

        task_name = str(row[task_name_idx]).strip() if row[task_name_idx] else ''
        if not task_name:
            continue

        # 获取 stage_id
        stage_id = None
        if stage_id_idx and row[stage_id_idx]:
            try:
                stage_id = int(row[stage_id_idx])
            except (ValueError, TypeError):
                # 尝试按名称查找
                stage_name = str(row[stage_id_idx]).strip()
                cursor.execute('SELECT stage_id FROM stage_map WHERE stage_name LIKE ?', (f'%{stage_name}%',))
                result = cursor.fetchone()
                if result:
                    stage_id = result[0]

        if not stage_id:
            continue

        seq = int(row[seq_idx]) if seq_idx and row[seq_idx] else 1
        days = int(row[days_idx]) if days_idx and row[days_idx] else 0
        critical = map_critical_path(str(row[critical_idx]).strip() if critical_idx and row[critical_idx] else '🟢')
        owner = str(row[owner_idx]).strip() if owner_idx and row[owner_idx] else '客户主导'
        description = str(row[desc_idx]).strip() if desc_idx and row[desc_idx] else ''
        acceptance = str(row[accept_idx]).strip() if accept_idx and row[accept_idx] else ''
        sort_order = int(row[sort_idx]) if sort_idx and row[sort_idx] else count + 1

        try:
            cursor.execute('''
                INSERT OR IGNORE INTO task_detail
                (stage_id, task_name, seq, default_days, critical_path, owner, description, acceptance_criteria, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (stage_id, task_name, seq, days, critical, owner, description, acceptance, sort_order))
            if cursor.rowcount > 0:
                count += 1
        except sqlite3.Error as e:
            print(f"   ⚠️ 任务 '{task_name}' 导入失败: {e}")

    return count


def import_pitfalls(cursor: sqlite3.Cursor, sheet) -> int:
    """导入避坑指南数据"""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0

    header = [str(h).strip() if h else '' for h in rows[0]]

    count = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue

        wrong_idx = find_column(header, ['wrong_action', '错误做法', '错误'])
        right_idx = find_column(header, ['right_action', '合规做法', '正确'])
        stage_idx = find_column(header, ['stage_ref', '关联阶段', '阶段'])
        impact_idx = find_column(header, ['impact_level', '影响等级', '影响'])
        source_idx = find_column(header, ['source', '来源'])

        if wrong_idx is None or right_idx is None:
            continue

        wrong_action = str(row[wrong_idx]).strip() if row[wrong_idx] else ''
        right_action = str(row[right_idx]).strip() if row[right_idx] else ''

        if not wrong_action or not right_action:
            continue

        stage_ref = str(row[stage_idx]).strip() if stage_idx and row[stage_idx] else ''
        impact = str(row[impact_idx]).strip() if impact_idx and row[impact_idx] else '中'
        source = str(row[source_idx]).strip() if source_idx and row[source_idx] else 'Excel导入'

        try:
            cursor.execute('''
                INSERT OR IGNORE INTO pitfall_guide
                (stage_ref, wrong_action, right_action, impact_level, source)
                VALUES (?, ?, ?, ?, ?)
            ''', (stage_ref, wrong_action, right_action, impact, source))
            if cursor.rowcount > 0:
                count += 1
        except sqlite3.Error as e:
            print(f"   ⚠️ 避坑 '{wrong_action[:20]}...' 导入失败: {e}")

    return count


def import_dependencies(cursor: sqlite3.Cursor, sheet) -> int:
    """导入任务依赖数据"""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0

    header = [str(h).strip() if h else '' for h in rows[0]]

    count = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue

        task_id_idx = find_column(header, ['task_id', '依赖方', '后续任务'])
        depends_idx = find_column(header, ['depends_on', '被依赖', '前置任务'])
        type_idx = find_column(header, ['dependency_type', '依赖类型', '类型'])

        if task_id_idx is None or depends_idx is None:
            continue

        try:
            task_id = int(row[task_id_idx]) if row[task_id_idx] else None
            depends_on = int(row[depends_idx]) if row[depends_idx] else None
        except (ValueError, TypeError):
            continue

        if not task_id or not depends_on or task_id == depends_on:
            continue

        dep_type = str(row[type_idx]).strip() if type_idx and row[type_idx] else '完成方可开始'

        try:
            cursor.execute('''
                INSERT OR IGNORE INTO task_dependency
                (task_id, depends_on, dependency_type)
                VALUES (?, ?, ?)
            ''', (task_id, depends_on, dep_type))
            if cursor.rowcount > 0:
                count += 1
        except sqlite3.Error:
            pass  # 忽略重复或约束错误

    return count


def find_column(header: List[str], names: List[str]) -> Optional[int]:
    """在表头中查找匹配的列索引"""
    for name in names:
        for i, h in enumerate(header):
            if name.lower() in h.lower():
                return i
    return None


def map_critical_path(value: str) -> str:
    """映射关键路径符号"""
    value = value.strip().upper()
    if value in ['🔴', 'RED', 'CRITICAL', '高', '关键', '1']:
        return '🔴'
    elif value in ['🟡', 'YELLOW', 'IMPORTANT', '中', '重要', '2']:
        return '🟡'
    else:
        return '🟢'


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='从 Excel 导入数据')
    parser.add_argument('--db', default='../data/innogreen_pmo.db', help='数据库路径')
    parser.add_argument('--excel', required=True, help='Excel 文件路径')
    args = parser.parse_args()

    success = import_from_excel(args.db, args.excel)
    sys.exit(0 if success else 1)
