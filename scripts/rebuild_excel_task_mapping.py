#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build composite-key mapping draft: (section, parent, leaf) → task_code."""
from __future__ import annotations

import csv
import re
import sqlite3
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "mappings"
OUT.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(ROOT / "Projects.xlsx", data_only=True, read_only=True)

# key: (section, parent_a, leaf_b) -> count, sheets
Key = tuple[str, str, str]
leaf_rows: dict[Key, dict] = {}

for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    status_col = next((i for i, h in enumerate(header) if h == "状态"), None)
    if status_col is None:
        status_col = next((i for i, h in enumerate(header) if "状态" in h), 3)

    section = ""
    last_parent = ""
    for row in rows[1:]:
        a = str(row[0]).replace("\n", " ").strip() if row and row[0] is not None else ""
        b = str(row[1]).replace("\n", " ").strip() if len(row) > 1 and row[1] is not None else ""
        st = ""
        if status_col is not None and len(row) > status_col and row[status_col] is not None:
            st = str(row[status_col]).strip()

        if a and not b and not st:
            section = a
            last_parent = ""
            continue

        if not b and not (a and st):
            continue

        # leaf row
        if a:
            last_parent = a
            parent = a
            leaf = b if b else a
            # if only A has value and it's the leaf (rare)
            if not b:
                parent = section
                leaf = a
        else:
            parent = last_parent
            leaf = b

        key = (section, parent, leaf)
        if key not in leaf_rows:
            leaf_rows[key] = {"count": 0, "sheets": set(), "statuses": Counter()}
        leaf_rows[key]["count"] += 1
        leaf_rows[key]["sheets"].add(sheet)
        if st:
            leaf_rows[key]["statuses"][st] += 1
wb.close()

conn = sqlite3.connect(ROOT / "data" / "innogreen_pmo.db")
conn.row_factory = sqlite3.Row
db_tasks = {
    r["task_code"]: dict(r)
    for r in conn.execute(
        """
        SELECT t.task_id, t.task_code, t.task_name, s.stage_name
        FROM task_detail t JOIN stage_map s ON s.stage_id=t.stage_id
        """
    )
}
# also by normalized name
def norm(s: str) -> str:
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"\s+", "", s)
    return s

by_name = {norm(v["task_name"]): v for v in db_tasks.values()}
conn.close()

# Manual composite rules: (parent_contains OR section_contains, leaf_equals_or_contains) -> task_code
# Order matters — first match wins
RULES: list[tuple[tuple[str, ...], tuple[str, ...], str, str]] = [
    # parent keywords, leaf keywords, task_code, note
    (("安评", "安全预评价", "反应安全"), ("化工反应", "反应安全风险", "反应风险评估"), "2.2.1", "安评-化工反应安全风险评估"),
    (("安评", "安全预评价"), ("报告编制", "编制定稿"), "2.2.2", "安评-报告编制"),
    (("安评", "安全预评价"), ("报告审批", "内部专家", "内部审批"), "2.2.3", "安评-报告审批"),
    (("安评", "安全", "安设"), ("专篇编制",), "2.2.4", "安设专篇编制"),
    (("安评", "安全", "安设"), ("专篇审批", "专篇备案", "内部专家"), "2.2.5", "安设专篇审批"),
    (("安评", "安全", "工艺审批"), ("安监处设计审查", "安全条件", "安监处进行"), "2.2.6", "安全条件/安监审查"),
    (("环评", "环境"), ("报告编制",), "2.3.1", "环评-报告编制"),
    (("环评", "环境"), ("报告审批", "内部审批"), "2.3.2", "环评-报告审批"),
    (("环评", "环境"), ("市局公示", "公示、审批", "公示审批"), "2.3.3", "环评-市局"),
    (("职评", "职业", "卫评", "职卫"), ("报告编制", "评价报告"), "2.4.1", "职评-报告编制"),
    (("职评", "职业", "卫评"), ("报告审批", "内部审批"), "2.4.2", "职评-报告审批"),
    (("职评", "职业", "卫评"), ("专篇编制",), "2.4.3", "职评专篇编制"),
    (("职评", "职业", "卫评"), ("专篇审批", "专篇备案"), "2.4.4", "职评专篇备案"),
    (("试生产",), ("方案编制", "编制定稿"), "6.1.1", "试生产方案编制"),
    (("试生产",), ("内部专家", "方案内部"), "6.1.2", "试生产方案评审"),
    (("试生产",), ("安全评审", "安监处进行开工前"), "6.1.3", "试生产安全评审"),
    (("试生产",), ("启动", "开始试生产"), "6.1.4", "启动试生产"),
    (("试生产",), ("试生产",), "6.1", "试生产父任务"),
    (("三同时", "竣工验收", "正式投用"), ("安全竣工", "安全设施竣工"), "6.2.1", "安全三同时验收"),
    (("三同时", "竣工验收"), ("环保竣工", "环保验收", "环境保护验收"), "6.2.2", "环保验收"),
    (("三同时", "竣工验收"), ("职业卫生", "职业病防护"), "6.2.3", "职卫验收"),
    (("供应链", "危化"), ("确认技术方案", "技术方案"), "3.8.2", "危化品技术方案"),
    (("供应链", "危化"), ("签订服务合同", "服务合同"), "3.8.3", "危化品服务合同"),
    (("供应链", "危化"), ("签订意向书", "意向书"), "3.8.1", "危化品意向书"),
    (("供应链", "危化"), ("正式提供服务",), "3.8.4", "危化品服务"),
    (("危废",), ("签订处置合同", "处置合同", "签订危废"), "3.6.2", "危废处置合同"),
    (("能源", "冷热", "蒸汽"), ("签订", "能源供应", "供能"), "3.7.3", "能源供应协议"),
]

SIMPLE_LEAF = {
    "项目预准入评估": "1.1",
    "项目准入评估": "1.2",
    "厂房租赁合同及HSE协议": "1.3",
    "厂房移交": "1.4",
    "项目投资信息报送": "1.5",
    "获取客户项目组成员信息": "2.1.1",
    "项目团队对接与沟通": "2.1.1",
    "项目可行性研究报告或初步设计": "2.1.2",
    "详细设计/施工图设计": "2.1.3",
    "施工图设计": "2.1.3",
    "市局公示、审批": "2.3.3",
    "建设工程项目信息报送": "4.1.1",
    "审图合同信息报送": "4.1.2",
    "设计合同信息报送": "4.1.3",
    "施工合同信息报送": "4.1.4",
    "监理合同信息报送": "4.1.5",
    "项目信息报送": "4.1.1",
    "合同信息报送": "4.1",
    "施工图审图": "4.2.1",
    "消防设计审查": "4.2.2",
    "防雷设计审查": "4.2.3",
    "施工许可证": "4.3.1",
    "安全生产管控平台 （承包商安全备案）": "4.3.2",
    "安全生产管控平台（承包商安全准入备案）": "4.3.2",
    "施工前准备工作（安全技术交底、临时设施搭建等）": "5.1",
    "施工前准备工作（安全技术交底）": "5.1.1",
    "装修土建结构施工": "5.2.1",
    "现场装修施工": "5.2.1",
    "机电工程施工": "5.3.1",
    "机电设备安装施工": "5.3.1",
    "设备及系统调试": "5.3.2",
    "竣工验收会": "5.4.1",
    "消防验收": "5.4.2",
    "防雷验收": "5.4.3",
    "综合竣工验收": "5.4.4",
    "安全竣工验收": "6.2.1",
    "环保验收": "6.2.2",
    "职业卫生验收": "6.2.3",
    "正式投用": "7.1",
    "签订意向书": "3.5.2",  # often sewage; may need context
    "安监处设计审查": "2.2.6",
    "化工反应安全风险评估": "2.2.1",
    "安监处进行开工前安全条件确认": "6.1.3",
    "专篇备案": "2.4.4",
}

# Utility service context → prefix mapping for 商务谈判/合同签订/正式提供服务
UTILITY_PARENT = {
    "咨询": ("3.1.1", "3.1.2", "3.1.3"),
    "物业": ("3.2.1", "3.2.2", "3.2.3"),
    "消防": ("3.3.1", "3.3.2", "3.3.3"),
    "环保": ("3.4.1", "3.4.2", "3.4.3"),
    "污水": ("3.5.1", "3.5.3", "3.5.4"),
    "危废": ("3.6.1", "3.6.2", "3.6.3"),
    "能源": ("3.7.1", "3.7.3", "3.7.4"),
    "冷热": ("3.7.1", "3.7.3", "3.7.4"),
    "蒸汽": ("3.7.1", "3.7.3", "3.7.4"),
    "危化": ("3.8.1", "3.8.3", "3.8.4"),
    "供应链": ("3.8.1", "3.8.3", "3.8.4"),
    "气体": ("3.9.1", "3.9.2", "3.9.3"),
}


def match_code(section: str, parent: str, leaf: str) -> tuple[str, str, str]:
    """Return (task_code, confidence, notes)."""
    # 1) structured rules first
    for pkeys, lkeys, code, note in RULES:
        if any(k in section or k in parent for k in pkeys) and any(k in leaf for k in lkeys):
            return code, "rule", note

    # 2) utility triad by parent type (before SIMPLE_LEAF — avoid 正式提供服务→7.1)
    if leaf in ("商务谈判", "合同签订", "正式提供服务", "签订意向书"):
        for uk, codes in UTILITY_PARENT.items():
            if uk in parent or uk in section:
                if leaf == "商务谈判":
                    return codes[0], "rule", f"公用工程-{uk}-谈判"
                if leaf in ("合同签订", "签订意向书"):
                    return codes[1], "rule", f"公用工程-{uk}-合同"
                if leaf == "正式提供服务":
                    return codes[2], "rule", f"公用工程-{uk}-服务"
        if leaf != "签订意向书":
            return "", "needs_context", "公用工程类同名任务，需标明服务类型（物业/消防/污水…）"

    if leaf in SIMPLE_LEAF:
        code = SIMPLE_LEAF[leaf]
        note = ""
        if leaf in ("正式提供服务", "商务谈判", "合同签订", "签订意向书"):
            note = "同名多义，默认映射；导入时建议按父级服务类型再校准"
        return code, "alias", note

    # exact name
    if norm(leaf) in by_name:
        t = by_name[norm(leaf)]
        return t["task_code"], "exact", ""

    # fuzzy containment
    best_code, best_sc = "", 0.0
    nl = norm(leaf)
    for name_n, t in by_name.items():
        if nl in name_n or name_n in nl:
            sc = min(len(nl), len(name_n)) / max(len(nl), len(name_n))
            if sc > best_sc:
                best_sc = sc
                best_code = t["task_code"]
    if best_sc >= 0.5:
        return best_code, "fuzzy", f"相似度{best_sc:.2f}"

    # skip / meta rows
    skip_kw = ("问题汇总", "应对依据", "备注", "附件")
    if any(k in leaf or k in section for k in skip_kw):
        return "", "skip", "非进度任务（问题/备注类）"

    return "", "unmapped", "待人工确认"


out_rows = []
for (section, parent, leaf), meta in sorted(
    leaf_rows.items(), key=lambda x: (-x[1]["count"], x[0][0], x[0][1], x[0][2])
):
    code, conf, notes = match_code(section, parent, leaf)
    t = db_tasks.get(code, {})
    out_rows.append(
        {
            "excel_section": section,
            "excel_parent": parent,
            "excel_task_name": leaf,
            "excel_count": meta["count"],
            "sheet_sample": ",".join(sorted(meta["sheets"])[:5]),
            "confidence": conf,
            "task_code": code,
            "task_id": t.get("task_id", ""),
            "task_name": t.get("task_name", ""),
            "stage_name": t.get("stage_name", ""),
            "notes": notes,
            "review_status": "todo" if conf in ("unmapped", "needs_context", "fuzzy") else "proposed",
        }
    )

fields = [
    "excel_section",
    "excel_parent",
    "excel_task_name",
    "excel_count",
    "sheet_sample",
    "confidence",
    "task_code",
    "task_id",
    "task_name",
    "stage_name",
    "notes",
    "review_status",
]
csv_path = OUT / "excel_task_to_task_code.csv"
with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=fields)
    w.writeheader()
    w.writerows(out_rows)

# stats
from collections import Counter as C

c = C(r["confidence"] for r in out_rows)
mapped = sum(1 for r in out_rows if r["task_code"])
print(f"composite_keys={len(out_rows)} mapped={mapped}")
print(dict(c))
print("wrote", csv_path)
