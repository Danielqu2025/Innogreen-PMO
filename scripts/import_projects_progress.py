#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Projects.xlsx + excel_task_to_task_code.reviewed.csv 导入企业进度与周记。

用法:
  # 默认 dry-run（不写库）
  python scripts/import_projects_progress.py

  # 试跑指定 sheet
  python scripts/import_projects_progress.py --sheets B5,A13

  # 真正写入（会先备份 DB）
  python scripts/import_projects_progress.py --apply
  python scripts/import_projects_progress.py --apply --sheets B5,A13
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import sqlite3
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "Projects.xlsx"
DEFAULT_MAP = ROOT / "data" / "mappings" / "excel_task_to_task_code.reviewed.csv"
DEFAULT_DB = ROOT / "data" / "innogreen_pmo.db"

STATUS_MAP = {
    "已完成": "已完成",
    "未启动": "待开始",
    "不涉及": "已跳过",
    "进行中（正常）": "进行中",
    "进行中（延后）": "进行中",
    "进行中": "进行中",
}

SKIP_STATUSES = frozenset({"", "应对依据"})


def log(msg: str) -> None:
    print(msg)


def load_mapping(path: Path) -> dict[tuple[str, str, str], dict]:
    """(section, parent, leaf) → row dict；仅 task_id 非空且非 skip。"""
    out: dict[tuple[str, str, str], dict] = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            conf = (r.get("confidence") or "").strip()
            if conf == "skip":
                continue
            tid = str(r.get("task_id") or "").strip()
            if not tid:
                continue
            key = (
                (r.get("excel_section") or "").strip(),
                (r.get("excel_parent") or "").strip(),
                (r.get("excel_task_name") or "").strip(),
            )
            if not key[2]:
                continue
            out[key] = r
    return out


def parse_cell_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).strftime("%Y-%m-%d %H:%M:%S")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return s


def parse_week_label(label: str, year_hint: int) -> tuple[str | None, int]:
    """
    解析「9月1日-5日」「7月13日-7月19日」「9月28日-30日」→ (week_start ISO, year_used)。
    year_hint 为当前推断年份；若月序相对上一列回绕则由调用方处理。
    """
    s = label.strip().replace(" ", "")
    # 去掉前缀如「每周工作进展（...）」
    m = re.search(
        r"(\d{1,2})月(\d{1,2})日(?:-(\d{1,2})月)?(?:-)?(\d{1,2})日?",
        s,
    )
    if not m:
        m2 = re.search(r"(\d{1,2})月(\d{1,2})日", s)
        if not m2:
            return None, year_hint
        month, day = int(m2.group(1)), int(m2.group(2))
        try:
            return date(year_hint, month, day).isoformat(), year_hint
        except ValueError:
            return None, year_hint
    month = int(m.group(1))
    day = int(m.group(2))
    try:
        return date(year_hint, month, day).isoformat(), year_hint
    except ValueError:
        return None, year_hint


def detect_columns(header: list[str]) -> dict:
    status_i = next((i for i, h in enumerate(header) if h == "状态"), None)
    vendor_i = next((i for i, h in enumerate(header) if "第三方" in h), None)
    plan_start_i = next((i for i, h in enumerate(header) if "计划开始" in h), None)
    plan_end_i = next((i for i, h in enumerate(header) if "计划完成" in h), None)
    actual_i = next(
        (i for i, h in enumerate(header) if "实际" in h and "完成" in h), None
    )
    week_cols: list[tuple[int, str]] = []
    start = (actual_i + 1) if actual_i is not None else (
        (status_i + 1) if status_i is not None else 7
    )
    for i in range(start, len(header)):
        h = header[i]
        if not h:
            continue
        if "月" in h and ("日" in h or "-" in h):
            week_cols.append((i, h))
    return {
        "status_i": status_i,
        "vendor_i": vendor_i,
        "plan_start_i": plan_start_i,
        "plan_end_i": plan_end_i,
        "actual_i": actual_i,
        "week_cols": week_cols,
    }


def iter_leaf_rows(rows: list, cols: dict):
    """Yield dicts for leaf progress rows (same logic as mapping rebuild)."""
    status_i = cols["status_i"]
    section = ""
    last_parent = ""
    for row in rows[1:]:
        a = str(row[0]).replace("\n", " ").strip() if row and row[0] is not None else ""
        b = (
            str(row[1]).replace("\n", " ").strip()
            if len(row) > 1 and row[1] is not None
            else ""
        )
        st = ""
        if status_i is not None and len(row) > status_i and row[status_i] is not None:
            st = str(row[status_i]).strip()

        if a and not b and not st:
            section = a
            last_parent = ""
            continue
        if not b and not (a and st):
            continue

        if a:
            last_parent = a
            parent = a
            leaf = b if b else a
            if not b:
                parent = section
                leaf = a
        else:
            parent = last_parent
            leaf = b

        yield {
            "section": section,
            "parent": parent,
            "leaf": leaf,
            "status_raw": st,
            "row": row,
        }


def ensure_project(conn: sqlite3.Connection, sheet: str, dry_run: bool) -> int | None:
    """project_code = sheet name；不存在则创建。返回 project_id。"""
    row = conn.execute(
        "SELECT project_id FROM project_profile WHERE project_code=?", (sheet,)
    ).fetchone()
    if row:
        return int(row[0])
    if dry_run:
        return None  # caller treats as would-create
    cur = conn.execute(
        """
        INSERT INTO project_profile (
          project_code, company_name, short_name, building,
          project_status, progress_percent, notes
        ) VALUES (?, ?, ?, ?, '进行中', 0, ?)
        """,
        (sheet, sheet, sheet, sheet, f"从 Projects.xlsx sheet={sheet} 导入"),
    )
    return int(cur.lastrowid)


def recalc_progress(conn: sqlite3.Connection, project_id: int) -> int:
    total = conn.execute(
        "SELECT COUNT(*) FROM task_detail WHERE is_active=1"
    ).fetchone()[0]
    if total == 0:
        pct = 0
    else:
        done = conn.execute(
            """
            SELECT COUNT(*) FROM project_progress p
            JOIN task_detail t ON t.task_id=p.task_id
            WHERE p.project_id=? AND p.status='已完成' AND t.is_active=1
            """,
            (project_id,),
        ).fetchone()[0]
        pct = min(100, max(0, round(100 * done / total)))

    blocked = conn.execute(
        """
        SELECT t.stage_id FROM project_progress p
        JOIN task_detail t ON t.task_id=p.task_id
        WHERE p.project_id=? AND p.status='卡点' AND t.is_active=1
        ORDER BY t.sort_order LIMIT 1
        """,
        (project_id,),
    ).fetchone()
    if blocked:
        status, stage_id = "卡点", blocked[0]
    else:
        inprog = conn.execute(
            """
            SELECT 1 FROM project_progress p
            JOIN task_detail t ON t.task_id=p.task_id
            WHERE p.project_id=? AND p.status='进行中' AND t.is_active=1 LIMIT 1
            """,
            (project_id,),
        ).fetchone()
        status = "进行中" if inprog else "进行中"
        stage_id = None
        # set current_stage to first 进行中 or first 待开始 with progress
        row = conn.execute(
            """
            SELECT t.stage_id FROM project_progress p
            JOIN task_detail t ON t.task_id=p.task_id
            WHERE p.project_id=? AND t.is_active=1
              AND p.status IN ('进行中','卡点','待开始')
            ORDER BY CASE p.status WHEN '卡点' THEN 0 WHEN '进行中' THEN 1 ELSE 2 END,
                     t.sort_order
            LIMIT 1
            """,
            (project_id,),
        ).fetchone()
        if row:
            stage_id = row[0]

    if stage_id is not None:
        conn.execute(
            """
            UPDATE project_profile
            SET progress_percent=?, project_status=?, current_stage_id=?
            WHERE project_id=?
            """,
            (pct, status, stage_id, project_id),
        )
    else:
        conn.execute(
            """
            UPDATE project_profile
            SET progress_percent=?, project_status=?
            WHERE project_id=?
            """,
            (pct, status, project_id),
        )
    return pct


def backup_db(db_path: Path) -> Path:
    bak_dir = ROOT / "data" / "backups"
    bak_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = bak_dir / f"innogreen_pmo_pre_import_{stamp}.db"
    shutil.copy2(db_path, dest)
    return dest


def build_week_starts(week_cols: list[tuple[int, str]]) -> list[tuple[int, str, str | None]]:
    """Assign year: headers chronological; start year 2025; bump on month rewind."""
    year = 2025
    prev_month: int | None = None
    out = []
    for idx, label in week_cols:
        m = re.search(r"(\d{1,2})月", label.replace(" ", ""))
        month = int(m.group(1)) if m else None
        if prev_month is not None and month is not None and month < prev_month - 2:
            # e.g. 12 → 1
            year += 1
        ws, year = parse_week_label(label, year)
        if month is not None:
            prev_month = month
        out.append((idx, label, ws))
    return out


def import_sheet(
    conn: sqlite3.Connection,
    sheet_name: str,
    ws,
    mapping: dict,
    dry_run: bool,
    stats: Counter,
) -> None:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log(f"  [{sheet_name}] 空表，跳过")
        return
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    cols = detect_columns(header)
    if cols["status_i"] is None:
        log(f"  [{sheet_name}] 无「状态」列，跳过")
        stats["sheets_skipped"] += 1
        return

    week_parsed = build_week_starts(cols["week_cols"])
    project_id = ensure_project(conn, sheet_name, dry_run=True)
    created_project = project_id is None
    if not dry_run:
        project_id = ensure_project(conn, sheet_name, dry_run=False)
    if created_project:
        stats["projects_created"] += 1
        log(f"  [{sheet_name}] {'将创建' if dry_run else '已创建'} project_code={sheet_name}")
    else:
        stats["projects_existing"] += 1

    for item in iter_leaf_rows(rows, cols):
        st_raw = item["status_raw"]
        if st_raw in SKIP_STATUSES:
            stats["rows_skip_status"] += 1
            continue
        key = (item["section"], item["parent"], item["leaf"])
        m = mapping.get(key)
        if not m:
            stats["rows_unmapped"] += 1
            continue
        task_id = int(m["task_id"])
        status = STATUS_MAP.get(st_raw)
        if not status:
            stats["rows_unknown_status"] += 1
            continue

        row = item["row"]
        vendor = None
        if cols["vendor_i"] is not None and len(row) > cols["vendor_i"]:
            v = row[cols["vendor_i"]]
            if v is not None and str(v).strip():
                vendor = str(v).strip()

        actual = None
        if cols["actual_i"] is not None and len(row) > cols["actual_i"]:
            actual = parse_cell_date(row[cols["actual_i"]])

        planned_start = None
        if cols["plan_start_i"] is not None and len(row) > cols["plan_start_i"]:
            planned_start = parse_cell_date(row[cols["plan_start_i"]])

        planned_end = None
        if cols["plan_end_i"] is not None and len(row) > cols["plan_end_i"]:
            planned_end = parse_cell_date(row[cols["plan_end_i"]])

        notes = "Excel状态:进行中（延后）" if st_raw == "进行中（延后）" else None

        # latest non-empty week note → blocker_note when 延后
        latest_note = None
        journal_entries: list[tuple[str, str, str]] = []  # week_start, label, note
        for col_i, label, week_start in week_parsed:
            if len(row) <= col_i:
                continue
            cell = row[col_i]
            if cell is None or not str(cell).strip():
                continue
            note = str(cell).strip()
            latest_note = note
            if week_start:
                journal_entries.append((week_start, label, note))

        blocker_note = None
        if st_raw == "进行中（延后）" and latest_note:
            blocker_note = latest_note[:2000]

        stats["rows_mapped"] += 1
        stats[f"status_{status}"] += 1
        stats["journal_cells"] += len(journal_entries)

        if dry_run or project_id is None:
            continue

        conn.execute(
            """
            INSERT INTO project_progress (
              project_id, task_id, status, assigned_to, started_at, completed_at,
              planned_start, planned_end, vendor, blocker_note, notes
            ) VALUES (?, ?, ?, NULL, NULL, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(project_id, task_id) DO UPDATE SET
              status=excluded.status,
              completed_at=excluded.completed_at,
              planned_start=excluded.planned_start,
              planned_end=excluded.planned_end,
              vendor=excluded.vendor,
              blocker_note=excluded.blocker_note,
              notes=excluded.notes,
              updated_at=datetime('now')
            """,
            (
                project_id,
                task_id,
                status,
                actual,
                planned_start,
                planned_end,
                vendor,
                blocker_note,
                notes,
            ),
        )
        stats["progress_upserts"] += 1

        for week_start, label, note in journal_entries:
            try:
                conn.execute(
                    """
                    INSERT INTO progress_journal (
                      project_id, task_id, week_start, week_label, note, source, actor
                    ) VALUES (?, ?, ?, ?, ?, 'excel_import', 'import')
                    """,
                    (project_id, task_id, week_start, label, note),
                )
                stats["journal_inserts"] += 1
            except sqlite3.IntegrityError:
                stats["journal_dupes"] += 1

    if not dry_run and project_id is not None:
        pct = recalc_progress(conn, project_id)
        stats["recalc_projects"] += 1
        log(f"  [{sheet_name}] progress_percent={pct}")


def main() -> int:
    ap = argparse.ArgumentParser(description="导入 Projects.xlsx 进度与周记")
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--mapping", type=Path, default=DEFAULT_MAP)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--sheets", type=str, default="", help="逗号分隔 sheet，空=全部")
    ap.add_argument("--apply", action="store_true", help="真正写库（默认 dry-run）")
    args = ap.parse_args()

    dry_run = not args.apply
    if not args.xlsx.exists():
        log(f"找不到 Excel: {args.xlsx}")
        return 1
    if not args.mapping.exists():
        log(f"找不到映射: {args.mapping}")
        return 1
    if not args.db.exists():
        log(f"找不到数据库: {args.db}")
        return 1

    mapping = load_mapping(args.mapping)
    log(f"映射键: {len(mapping)}（reviewed 且有 task_id）")
    log(f"模式: {'APPLY 写库' if args.apply else 'DRY-RUN 只统计'}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    want = [s.strip() for s in args.sheets.split(",") if s.strip()]
    sheets = want if want else list(wb.sheetnames)
    for s in sheets:
        if s not in wb.sheetnames:
            log(f"sheet 不存在: {s}")
            wb.close()
            return 1

    if args.apply:
        bak = backup_db(args.db)
        log(f"已备份: {bak}")

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA foreign_keys=ON")
    # ensure journal table
    journal_sql = ROOT / "sql" / "progress_journal.sql"
    if journal_sql.exists():
        conn.executescript(journal_sql.read_text(encoding="utf-8"))

    # ensure is_active + schedule columns
    tcols = [r[1] for r in conn.execute("PRAGMA table_info(task_detail)")]
    if "is_active" not in tcols:
        conn.execute(
            "ALTER TABLE task_detail ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1"
        )
    pcols = [r[1] for r in conn.execute("PRAGMA table_info(project_progress)")]
    for col in ("planned_start", "planned_end", "vendor"):
        if col not in pcols:
            conn.execute(f"ALTER TABLE project_progress ADD COLUMN {col} TEXT")
            log(f"已添加列 project_progress.{col}")

    stats: Counter = Counter()
    try:
        for name in sheets:
            log(f"处理 sheet: {name}")
            import_sheet(conn, name, wb[name], mapping, dry_run, stats)
            stats["sheets"] += 1
        if args.apply:
            conn.commit()
        else:
            conn.rollback()
    finally:
        wb.close()
        conn.close()

    log("\n=== 汇总 ===")
    for k in sorted(stats):
        log(f"  {k}: {stats[k]}")
    if dry_run:
        log("\n（dry-run 未写库。确认后加 --apply）")
    return 0


if __name__ == "__main__":
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    raise SystemExit(main())
