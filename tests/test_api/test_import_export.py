"""Smoke tests for Excel / DB export / import (roles + sheet filter)."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook


def test_viewer_cannot_export(viewer_client):
    r = viewer_client.get("/api/ops/export/excel")
    assert r.status_code == 403


def test_operator_can_export_excel(operator_client):
    r = operator_client.get("/api/ops/export/excel")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")


def test_operator_cannot_export_db(operator_client):
    """整库 .db 导出仅管理员；操作员应 403。"""
    r = operator_client.get("/api/ops/export/db")
    assert r.status_code == 403


def test_import_excel_rejects_oversized(operator_client, monkeypatch):
    """Excel 上传超限 → 413。"""
    import routers.ops as ops_mod

    monkeypatch.setattr(ops_mod, "MAX_EXCEL_UPLOAD_BYTES", 64)
    r = operator_client.post(
        "/api/ops/import/excel",
        files={
            "file": (
                "big.xlsx",
                b"x" * 200,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        params={"dry_run": True},
    )
    assert r.status_code == 413, r.text
    assert r.json()["detail"]["code"] == "ERR_PAYLOAD_TOO_LARGE"


def test_import_db_rejects_oversized(admin_client, monkeypatch):
    """DB 上传超限 → 413。"""
    import routers.ops as ops_mod

    monkeypatch.setattr(ops_mod, "MAX_DB_UPLOAD_BYTES", 64)
    r = admin_client.post(
        "/api/ops/import/db",
        files={"file": ("big.db", b"x" * 200, "application/x-sqlite3")},
        params={"confirm": True},
    )
    assert r.status_code == 413, r.text
    assert r.json()["detail"]["code"] == "ERR_PAYLOAD_TOO_LARGE"


def test_operator_can_import_excel_dry_run(operator_client):
    wb = Workbook()
    ws = wb.active
    ws.title = "企业档案"
    ws.append(["project_code", "company_name"])
    buf = BytesIO()
    wb.save(buf)
    r = operator_client.post(
        "/api/ops/import/excel",
        files={
            "file": (
                "t.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        params={"dry_run": True},
    )
    assert r.status_code == 200, r.text


def test_operator_can_download_template(operator_client):
    r = operator_client.get("/api/ops/import/template.xlsx")
    assert r.status_code == 200, r.text


def test_operator_cannot_import_db(operator_client):
    r = operator_client.post(
        "/api/ops/import/db",
        files={"file": ("t.db", b"SQLite format 3\x00", "application/x-sqlite3")},
        params={"confirm": True},
    )
    assert r.status_code == 403


def test_admin_export_xlsx(admin_client):
    r = admin_client.get("/api/ops/export/excel")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(BytesIO(r.content), read_only=True)
    names = set(wb.sheetnames)
    assert "阶段定义" in names
    assert "任务明细" in names
    assert "企业档案" in names
    assert "任务进度" in names
    assert "避坑指南" in names


def test_export_xlsx_sheets_filter(admin_client):
    r = admin_client.get(
        "/api/ops/export/excel",
        params={"sheets": "projects,progress"},
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content), read_only=True)
    assert set(wb.sheetnames) == {"企业档案", "任务进度"}


def test_export_xlsx_sheets_chinese_names(admin_client):
    r = admin_client.get(
        "/api/ops/export/excel",
        params={"sheets": "阶段定义,避坑指南"},
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content), read_only=True)
    assert set(wb.sheetnames) == {"阶段定义", "避坑指南"}


def test_export_xlsx_sheets_invalid(admin_client):
    r = admin_client.get(
        "/api/ops/export/excel",
        params={"sheets": "projects,nope"},
    )
    assert r.status_code == 400
    assert "sheets" in r.json()["detail"]["message"].lower() or "无效" in r.json()[
        "detail"
    ]["message"]


def test_admin_export_db_sqlite_header(admin_client):
    r = admin_client.get("/api/ops/export/db")
    assert r.status_code == 200, r.text
    assert r.content.startswith(b"SQLite format 3\x00")
    cd = r.headers.get("content-disposition", "")
    assert ".db" in cd


def test_admin_import_template_xlsx(admin_client):
    r = admin_client.get("/api/ops/import/template.xlsx")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(BytesIO(r.content), read_only=True)
    names = set(wb.sheetnames)
    assert "说明" in names
    assert "企业档案" in names
    assert "任务进度" in names
    guide = wb["说明"]
    guide_rows = list(guide.iter_rows(values_only=True))
    # 字段定义表表头
    assert any(
        row and row[0] == "字段" and row[1] == "是否必填" for row in guide_rows
    )
    assert any(row and row[0] == "project_code" for row in guide_rows)
    assert any(row and row[0] == "status" for row in guide_rows)
    # 可选值含任务状态
    flat = " ".join(
        str(c) for row in guide_rows for c in (row or ()) if c is not None
    )
    assert "待开始" in flat and "卡点" in flat
    assert "新建企业" in flat or "更新已有" in flat

    proj = wb["企业档案"]
    rows = list(proj.iter_rows(values_only=True))
    assert rows[0][0] == "project_code"
    assert len(rows) >= 2  # header + examples


def test_admin_import_db_requires_confirm(admin_client):
    snap = admin_client.get("/api/ops/export/db")
    assert snap.status_code == 200
    r = admin_client.post(
        "/api/ops/import/db",
        files={
            "file": ("roundtrip.db", snap.content, "application/x-sqlite3"),
        },
        params={"confirm": False},
    )
    assert r.status_code == 400
    assert "confirm" in r.json()["detail"]["message"].lower() or "确认" in r.json()[
        "detail"
    ]["message"]


def test_admin_import_db_roundtrip(admin_client):
    """用当前库快照再导入（全量替换），验证成功且 API 仍可用。"""
    snap = admin_client.get("/api/ops/export/db")
    assert snap.status_code == 200
    assert snap.content.startswith(b"SQLite format 3\x00")

    r = admin_client.post(
        "/api/ops/import/db",
        files={
            "file": ("roundtrip.db", snap.content, "application/x-sqlite3"),
        },
        params={"confirm": True},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    # 仅相对路径（backups/文件名），不暴露绝对路径
    bp = body["backup_path"].replace("\\", "/")
    assert bp.startswith("backups/")
    assert ":" not in bp  # 无盘符
    assert ".." not in bp
    assert bp.count("/") == 1

    # 替换后连接池应能重新打开；重新登录因用户表仍在
    health = admin_client.get("/health")
    assert health.status_code == 200
    projects = admin_client.get("/api/ops/projects")
    assert projects.status_code == 200
    assert isinstance(projects.json(), list)


def test_admin_import_db_rejects_non_sqlite(admin_client):
    r = admin_client.post(
        "/api/ops/import/db",
        files={"file": ("bad.db", b"not a database", "application/octet-stream")},
        params={"confirm": True},
    )
    assert r.status_code == 400


def test_admin_import_dry_run_and_apply(admin_client):
    # 取一个已有任务编号
    tasks = admin_client.get("/api/ops/tasks", params={"stage_id": 1}).json()
    assert tasks
    task_code = next(t["task_code"] for t in tasks if t.get("task_code"))

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    proj = wb.create_sheet("企业档案")
    proj.append(
        [
            "project_code",
            "company_name",
            "short_name",
            "business_type",
            "building",
            "project_status",
            "notes",
        ]
    )
    proj.append(
        [
            "ENT-TEST-XLSX",
            "导入测试企业",
            "测试企",
            "测试",
            "A栋",
            "进行中",
            "from import test",
        ]
    )

    prog = wb.create_sheet("任务进度")
    prog.append(["project_code", "task_code", "status", "assigned_to"])
    prog.append(["ENT-TEST-XLSX", task_code, "进行中", "pytest"])

    # 目录 sheet 应被忽略
    stages = wb.create_sheet("阶段定义")
    stages.append(["stage_id", "stage_name"])
    stages.append([999, "不应导入"])

    buf = BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    dry = admin_client.post(
        "/api/ops/import/excel",
        files={
            "file": (
                "import.xlsx",
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        params={"dry_run": True},
    )
    assert dry.status_code == 200, dry.text
    body = dry.json()
    assert body["dry_run"] is True
    assert body["projects_created"] == 1
    assert body["progress_upserted"] == 1
    assert any("忽略" in w for w in body["warnings"])

    # dry-run 不应写库
    listed = admin_client.get("/api/ops/projects").json()
    assert all(p["project_code"] != "ENT-TEST-XLSX" for p in listed)

    applied = admin_client.post(
        "/api/ops/import/excel",
        files={
            "file": (
                "import.xlsx",
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        params={"dry_run": False},
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["dry_run"] is False
    assert applied.json()["projects_created"] == 1

    listed2 = admin_client.get("/api/ops/projects").json()
    match = [p for p in listed2 if p["project_code"] == "ENT-TEST-XLSX"]
    assert len(match) == 1
    assert match[0]["company_name"] == "导入测试企业"

    progress = admin_client.get(
        f"/api/ops/projects/{match[0]['project_id']}/progress"
    ).json()
    hit = [p for p in progress if p["task_code"] == task_code]
    assert hit and hit[0]["status"] == "进行中"
