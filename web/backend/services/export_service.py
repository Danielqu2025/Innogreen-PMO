"""数据导出服务 — 生成多 sheet Excel（.xlsx）备份。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from models import PitfallGuide, ProjectProgress, ProjectProfile, StageMap, TaskDetail
from services.progress_service import VALID_STATUSES
from services.project_service import VALID_PROJECT_STATUSES

SHEET_STAGES = "阶段定义"
SHEET_TASKS = "任务明细"
SHEET_PROJECTS = "企业档案"
SHEET_PROGRESS = "任务进度"
SHEET_PITFALLS = "避坑指南"

# 导出可选 sheet：API key → 中文 sheet 名
EXPORT_SHEET_KEYS: dict[str, str] = {
    "stages": SHEET_STAGES,
    "tasks": SHEET_TASKS,
    "projects": SHEET_PROJECTS,
    "progress": SHEET_PROGRESS,
    "pitfalls": SHEET_PITFALLS,
}
# 亦接受中文名作为入参
_EXPORT_ALIAS: dict[str, str] = {
    **{k: k for k in EXPORT_SHEET_KEYS},
    **{v: k for k, v in EXPORT_SHEET_KEYS.items()},
}
ALL_EXPORT_KEYS: tuple[str, ...] = tuple(EXPORT_SHEET_KEYS.keys())

# 业务类型建议值（与前端表单一致；导入不强制枚举）
BUSINESS_TYPE_HINTS = ("研发小试", "中试", "研发小试和中试混合")


def parse_export_sheets(raw: str | list[str] | None) -> list[str]:
    """解析 sheets 参数，返回合法 API key 列表（去重、保序）。

    支持：逗号分隔字符串、列表；key 或中文 sheet 名。
    None / 空 → 全部 sheet。
    非法值 → ValueError。
    """
    if raw is None:
        return list(ALL_EXPORT_KEYS)
    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(",") if p.strip()]
    else:
        parts = [str(p).strip() for p in raw if str(p).strip()]
    if not parts:
        return list(ALL_EXPORT_KEYS)

    out: list[str] = []
    seen: set[str] = set()
    unknown: list[str] = []
    for p in parts:
        key = _EXPORT_ALIAS.get(p) or _EXPORT_ALIAS.get(p.lower())
        if key is None:
            unknown.append(p)
            continue
        if key not in seen:
            seen.add(key)
            out.append(key)
    if unknown:
        allowed = ", ".join(ALL_EXPORT_KEYS) + " 或 " + "、".join(EXPORT_SHEET_KEYS.values())
        raise ValueError(f"无效的 sheets: {', '.join(unknown)}；可选: {allowed}")
    return out


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build_export_workbook(
    db: Session,
    sheets: list[str] | None = None,
) -> bytes:
    """导出多 sheet xlsx。sheets 为 API key 列表；None 表示全部。"""
    keys = sheets if sheets is not None else list(ALL_EXPORT_KEYS)
    if not keys:
        keys = list(ALL_EXPORT_KEYS)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    if "stages" in keys:
        stages = db.execute(select(StageMap).order_by(StageMap.sort_order)).scalars().all()
        _write_sheet(
            wb,
            SHEET_STAGES,
            [
                "stage_id",
                "stage_name",
                "primary_owner",
                "critical_path",
                "default_days",
                "description",
                "sort_order",
            ],
            [
                [
                    s.stage_id,
                    s.stage_name,
                    s.primary_owner,
                    s.critical_path,
                    s.default_days,
                    s.description,
                    s.sort_order,
                ]
                for s in stages
            ],
        )

    if "tasks" in keys:
        tasks = db.execute(
            select(TaskDetail).order_by(TaskDetail.stage_id, TaskDetail.sort_order)
        ).scalars().all()
        _write_sheet(
            wb,
            SHEET_TASKS,
            [
                "task_id",
                "stage_id",
                "task_code",
                "task_name",
                "seq",
                "default_days",
                "critical_path",
                "owner",
                "description",
                "sort_order",
                "is_active",
            ],
            [
                [
                    t.task_id,
                    t.stage_id,
                    t.task_code,
                    t.task_name,
                    t.seq,
                    t.default_days,
                    t.critical_path,
                    t.owner,
                    t.description,
                    t.sort_order,
                    t.is_active,
                ]
                for t in tasks
            ],
        )

    if "projects" in keys:
        projects = db.execute(
            select(ProjectProfile)
            .options(joinedload(ProjectProfile.current_stage))
            .order_by(ProjectProfile.project_id)
        ).scalars().unique().all()
        _write_sheet(
            wb,
            SHEET_PROJECTS,
            [
                "project_code",
                "company_name",
                "short_name",
                "full_name",
                "business_type",
                "building",
                "current_stage_id",
                "project_status",
                "progress_percent",
                "notes",
            ],
            [
                [
                    p.project_code,
                    p.company_name,
                    p.short_name,
                    p.full_name,
                    p.business_type,
                    p.building,
                    p.current_stage_id,
                    p.project_status,
                    p.progress_percent,
                    p.notes,
                ]
                for p in projects
            ],
        )

    if "progress" in keys:
        progress_rows = db.execute(
            select(ProjectProgress, ProjectProfile, TaskDetail)
            .join(ProjectProfile, ProjectProgress.project_id == ProjectProfile.project_id)
            .join(TaskDetail, ProjectProgress.task_id == TaskDetail.task_id)
            .order_by(ProjectProfile.project_code, TaskDetail.task_code)
        ).all()
        _write_sheet(
            wb,
            SHEET_PROGRESS,
            [
                "project_code",
                "task_code",
                "task_name",
                "status",
                "assigned_to",
                "planned_start",
                "planned_end",
                "started_at",
                "completed_at",
                "vendor",
                "blocker_note",
                "notes",
            ],
            [
                [
                    proj.project_code,
                    task.task_code,
                    task.task_name,
                    prog.status,
                    prog.assigned_to,
                    prog.planned_start,
                    prog.planned_end,
                    prog.started_at,
                    prog.completed_at,
                    prog.vendor,
                    prog.blocker_note,
                    prog.notes,
                ]
                for prog, proj, task in progress_rows
            ],
        )

    if "pitfalls" in keys:
        pitfalls = db.execute(
            select(PitfallGuide).order_by(PitfallGuide.pitfall_id)
        ).scalars().all()
        _write_sheet(
            wb,
            SHEET_PITFALLS,
            [
                "pitfall_id",
                "stage_ref",
                "wrong_action",
                "right_action",
                "standard_ref",
                "impact_level",
                "error_index",
                "trigger_condition",
                "remediation",
                "notes",
                "source",
                "verified",
            ],
            [
                [
                    p.pitfall_id,
                    p.stage_ref,
                    p.wrong_action,
                    p.right_action,
                    p.standard_ref,
                    p.impact_level,
                    p.error_index,
                    p.trigger_condition,
                    p.remediation,
                    p.notes,
                    p.source,
                    p.verified,
                ]
                for p in pitfalls
            ],
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _guide_table_rows() -> list[list]:
    """说明 sheet：字段定义表（字段 | 是否必填 | 说明 | 可选值）。"""
    status_opts = " / ".join(sorted(VALID_PROJECT_STATUSES))
    task_status_opts = " / ".join(sorted(VALID_STATUSES))
    biz_opts = " / ".join(BUSINESS_TYPE_HINTS) + "（建议值，亦可自填）"

    rows: list[list] = [
        ["字段", "是否必填", "说明", "可选值"],
        [],
        ["【使用说明】", "", "", ""],
        [
            "新建企业",
            "",
            "在「企业档案」填一行：使用尚未存在的 project_code + company_name。"
            "可同文件在「任务进度」用同一 project_code 写初始进度。",
            "",
        ],
        [
            "更新已有企业进度",
            "",
            "「企业档案」可省略或只改需要更新的字段；"
            "在「任务进度」用已有 project_code + 系统已有 task_code 更新状态。",
            "",
        ],
        [
            "导入范围",
            "",
            "仅写入「企业档案」「任务进度」；「说明」及其他 sheet 会被忽略。"
            "建议先试跑再正式导入。",
            "",
        ],
        [
            "日期格式",
            "",
            "建议 YYYY-MM-DD（如 2026-01-15）",
            "",
        ],
        [],
        ["【企业档案】", "", "", ""],
        ["project_code", "必填", "企业编号，唯一标识；已存在则更新，不存在则新建", "如 ENT-04"],
        ["company_name", "必填（新建）", "企业显示名称；新建必填，更新时可留空表示不改", ""],
        ["short_name", "可选", "企业简称", ""],
        ["full_name", "可选", "企业全称", ""],
        ["business_type", "可选", "业务类型", biz_opts],
        ["building", "可选", "所在楼栋/工位", "如 A栋、F6d-1"],
        [
            "current_stage_id",
            "可选",
            "当前阶段 ID（整数）；亦可由任务进度自动推算",
            "0–8（系统阶段）",
        ],
        ["project_status", "可选", "项目整体状态；缺省新建为「未开始」", status_opts],
        ["notes", "可选", "备注", ""],
        [],
        ["【任务进度】", "", "", ""],
        ["project_code", "必填", "对应企业编号（须已存在，或同文件企业档案中新建）", ""],
        [
            "task_code",
            "必填",
            "任务编号，须为系统中已启用的任务（如 1.1、2.1.3）",
            "见「任务清单」或导出「任务明细」",
        ],
        ["status", "必填", "该任务进度状态", task_status_opts],
        ["assigned_to", "可选", "负责人", ""],
        ["planned_start", "可选", "计划开始日", "YYYY-MM-DD"],
        ["planned_end", "可选", "计划结束日", "YYYY-MM-DD"],
        ["started_at", "可选", "实际开始日", "YYYY-MM-DD"],
        ["completed_at", "可选", "实际完成日", "YYYY-MM-DD"],
        ["vendor", "可选", "供应商/服务方", ""],
        ["blocker_note", "可选", "卡点说明（status=卡点时建议填写）", ""],
        ["notes", "可选", "备注", ""],
        ["task_name", "忽略", "导出时附带；导入时忽略，以系统任务名为准", ""],
    ]
    return rows


def build_import_template() -> bytes:
    """导入用空白模板（说明字段表 + 企业档案/任务进度示例）。"""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    guide = wb.create_sheet("说明", 0)
    guide.append(["Innogreen PMO — Excel 导入模板（填写指南）"])
    guide.append([])
    for row in _guide_table_rows():
        guide.append(row)

    projects = wb.create_sheet(SHEET_PROJECTS)
    projects.append(
        [
            "project_code",
            "company_name",
            "short_name",
            "full_name",
            "business_type",
            "building",
            "current_stage_id",
            "project_status",
            "notes",
        ]
    )
    projects.append(
        [
            "ENT-NEW-01",
            "示例新材料有限公司",
            "示例新材料",
            "示例新材料（上海）有限公司",
            "研发小试",
            "A栋",
            1,
            "未开始",
            "模板示例：新建企业，导入前请修改编号与名称",
        ]
    )
    projects.append(
        [
            "ENT-NEW-02",
            "示例生物科技有限公司",
            "示例生科",
            "",
            "中试",
            "B栋",
            2,
            "进行中",
            "模板示例：可删除本行",
        ]
    )

    progress = wb.create_sheet(SHEET_PROGRESS)
    progress.append(
        [
            "project_code",
            "task_code",
            "status",
            "assigned_to",
            "planned_start",
            "planned_end",
            "started_at",
            "completed_at",
            "vendor",
            "blocker_note",
            "notes",
        ]
    )
    progress.append(
        [
            "ENT-01",
            "1.1",
            "进行中",
            "张三",
            "2026-01-01",
            "2026-01-15",
            "2026-01-02",
            "",
            "",
            "",
            "模板示例：更新已有企业 ENT-01 的任务进度（请改成真实 task_code）",
        ]
    )
    progress.append(
        [
            "ENT-NEW-01",
            "1.1",
            "待开始",
            "李四",
            "2026-02-01",
            "2026-02-10",
            "",
            "",
            "",
            "",
            "模板示例：可与上方「企业档案」新建行配套使用",
        ]
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
