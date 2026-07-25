"""数据导入服务 — 按导出格式安全 upsert 企业档案、任务进度与避坑指南。

避坑指南仅支持新增（pitfall_id 空时新建；已有则覆盖关键字段但不删除）；
阶段/任务目录不支持导入（全量重导风险高），这些 sheet 若存在会被忽略。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from models import PitfallGuide, ProjectProfile, ProjectProgress, StageMap, TaskDetail
from services.audit import log_action
from services.export_service import SHEET_PITFALLS, SHEET_PROGRESS, SHEET_PROJECTS
from services.pitfall_service import _task_level
from services.progress_service import (
    VALID_STATUSES,
    recalculate_progress_percent,
    sync_project_status,
)
from services.project_service import VALID_PROJECT_STATUSES

# 兼容英文 sheet 名
_PROJECT_SHEETS = frozenset({SHEET_PROJECTS, "projects", "Projects"})
_PROGRESS_SHEETS = frozenset({SHEET_PROGRESS, "progress", "Progress"})
_PITFALL_SHEETS = frozenset({SHEET_PITFALLS, "pitfalls", "Pitfalls"})


@dataclass
class ImportSummary:
    dry_run: bool
    projects_created: int = 0
    projects_updated: int = 0
    progress_upserted: int = 0
    progress_skipped: int = 0
    pitfalls_created: int = 0
    pitfalls_updated: int = 0
    pitfalls_skipped: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "dry_run": self.dry_run,
            "projects_created": self.projects_created,
            "projects_updated": self.projects_updated,
            "progress_upserted": self.progress_upserted,
            "progress_skipped": self.progress_skipped,
            "pitfalls_created": self.pitfalls_created,
            "pitfalls_updated": self.pitfalls_updated,
            "pitfalls_skipped": self.pitfalls_skipped,
            "warnings": self.warnings,
            "errors": self.errors,
        }


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _header_map(row: tuple) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(row):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key:
            out[key] = i
    return out


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        i = headers.get(n.lower())
        if i is not None:
            return i
    return None


def _find_sheet(wb, aliases: frozenset[str]):
    for name in wb.sheetnames:
        if name in aliases:
            return wb[name]
    return None


def import_projects_progress_xlsx(
    db: Session,
    file_bytes: bytes,
    actor: str,
    *,
    dry_run: bool = True,
    ip_address: str | None = None,
    user_agent: str | None = None,
) -> ImportSummary:
    summary = ImportSummary(dry_run=dry_run)

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        summary.errors.append(f"无法读取 Excel: {e}")
        return summary

    project_ws = _find_sheet(wb, _PROJECT_SHEETS)
    progress_ws = _find_sheet(wb, _PROGRESS_SHEETS)
    pitfall_ws = _find_sheet(wb, _PITFALL_SHEETS)

    if project_ws is None and progress_ws is None and pitfall_ws is None:
        summary.errors.append(
            f"未找到可用 sheet（需要「{SHEET_PROJECTS}」、「{SHEET_PROGRESS}」或「{SHEET_PITFALLS}」）"
        )
        return summary

    # 只读提示：目录类 sheet 不会写入
    ignored = [
        n
        for n in wb.sheetnames
        if n not in _PROJECT_SHEETS
        and n not in _PROGRESS_SHEETS
        and n not in _PITFALL_SHEETS
    ]
    if ignored:
        summary.warnings.append(
            "以下 sheet 已忽略（本接口不导入阶段/任务目录）: "
            + ", ".join(ignored)
        )

    # task_code → TaskDetail（启用任务）
    tasks = db.execute(
        select(TaskDetail).where(TaskDetail.is_active == 1)
    ).scalars().all()
    by_code: dict[str, TaskDetail] = {}
    for t in tasks:
        if t.task_code:
            by_code[t.task_code.strip()] = t

    projects_by_code: dict[str, ProjectProfile] = {
        p.project_code: p
        for p in db.execute(select(ProjectProfile)).scalars().all()
    }
    touched_project_ids: set[int] = set()

    if project_ws is not None:
        rows = list(project_ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = _header_map(rows[0])
            code_i = _col(headers, "project_code", "企业编号", "编号")
            name_i = _col(headers, "company_name", "企业名称", "名称")
            if code_i is None or name_i is None:
                summary.errors.append(
                    f"「{SHEET_PROJECTS}」缺少 project_code / company_name 列"
                )
            else:
                short_i = _col(headers, "short_name", "简称")
                full_i = _col(headers, "full_name", "全称")
                biz_i = _col(headers, "business_type", "业务类型")
                bld_i = _col(headers, "building", "楼栋")
                status_i = _col(headers, "project_status", "项目状态", "状态")
                notes_i = _col(headers, "notes", "备注")
                stage_i = _col(headers, "current_stage_id", "当前阶段id", "阶段id")

                for ridx, row in enumerate(rows[1:], start=2):
                    if not row:
                        continue
                    code = _cell_str(row[code_i] if code_i < len(row) else None)
                    name = _cell_str(row[name_i] if name_i < len(row) else None)
                    if not code or not name:
                        continue

                    short = _cell_str(row[short_i]) if short_i is not None else None
                    full = _cell_str(row[full_i]) if full_i is not None else None
                    biz = _cell_str(row[biz_i]) if biz_i is not None else None
                    bld = _cell_str(row[bld_i]) if bld_i is not None else None
                    notes = _cell_str(row[notes_i]) if notes_i is not None else None
                    status = _cell_str(row[status_i]) if status_i is not None else None
                    stage_raw = row[stage_i] if stage_i is not None and stage_i < len(row) else None
                    stage_id = None
                    if stage_raw is not None and str(stage_raw).strip():
                        try:
                            stage_id = int(stage_raw)
                        except (TypeError, ValueError):
                            summary.warnings.append(
                                f"第 {ridx} 行 current_stage_id 无效: {stage_raw}"
                            )

                    if status and status not in VALID_PROJECT_STATUSES:
                        summary.warnings.append(
                            f"第 {ridx} 行项目状态无效已忽略: {status}"
                        )
                        status = None

                    existing = projects_by_code.get(code)
                    if existing is None:
                        summary.projects_created += 1
                        if dry_run:
                            # 试跑占位，便于同文件内进度行按编号匹配
                            stub = ProjectProfile(
                                project_id=-len(projects_by_code) - 1,
                                project_code=code,
                                company_name=name,
                            )
                            projects_by_code[code] = stub
                        else:
                            p = ProjectProfile(
                                project_code=code,
                                company_name=name,
                                short_name=short or code,
                                full_name=full,
                                business_type=biz,
                                building=bld,
                                current_stage_id=stage_id,
                                project_status=status or "未开始",
                                progress_percent=0,
                                notes=notes,
                            )
                            db.add(p)
                            db.flush()
                            projects_by_code[code] = p
                            touched_project_ids.add(p.project_id)
                    else:
                        summary.projects_updated += 1
                        if not dry_run:
                            existing.company_name = name
                            if short is not None:
                                existing.short_name = short
                            if full is not None:
                                existing.full_name = full
                            if biz is not None:
                                existing.business_type = biz
                            if bld is not None:
                                existing.building = bld
                            if notes is not None:
                                existing.notes = notes
                            if status is not None:
                                existing.project_status = status
                            if stage_id is not None:
                                existing.current_stage_id = stage_id
                            touched_project_ids.add(existing.project_id)

    if progress_ws is not None:
        rows = list(progress_ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = _header_map(rows[0])
            code_i = _col(headers, "project_code", "企业编号")
            task_i = _col(headers, "task_code", "任务编号")
            status_i = _col(headers, "status", "状态")
            if code_i is None or task_i is None or status_i is None:
                summary.errors.append(
                    f"「{SHEET_PROGRESS}」缺少 project_code / task_code / status 列"
                )
            else:
                assigned_i = _col(headers, "assigned_to", "负责人")
                plan_s_i = _col(headers, "planned_start", "计划开始")
                plan_e_i = _col(headers, "planned_end", "计划完成")
                start_i = _col(headers, "started_at", "实际开始")
                done_i = _col(headers, "completed_at", "实际完成")
                vendor_i = _col(headers, "vendor", "第三方")
                blocker_i = _col(headers, "blocker_note", "卡点说明")
                notes_i = _col(headers, "notes", "备注")

                for ridx, row in enumerate(rows[1:], start=2):
                    if not row:
                        continue
                    pcode = _cell_str(row[code_i] if code_i < len(row) else None)
                    tcode = _cell_str(row[task_i] if task_i < len(row) else None)
                    status = _cell_str(row[status_i] if status_i < len(row) else None)
                    if not pcode or not tcode or not status:
                        continue
                    if status not in VALID_STATUSES:
                        summary.progress_skipped += 1
                        summary.warnings.append(
                            f"进度第 {ridx} 行状态无效: {status}"
                        )
                        continue

                    project = projects_by_code.get(pcode)
                    if project is None:
                        summary.progress_skipped += 1
                        summary.warnings.append(
                            f"进度第 {ridx} 行未知企业编号: {pcode}"
                        )
                        continue

                    task = by_code.get(tcode)
                    if task is None:
                        summary.progress_skipped += 1
                        summary.warnings.append(
                            f"进度第 {ridx} 行未知或已停用任务编号: {tcode}"
                        )
                        continue

                    summary.progress_upserted += 1
                    if dry_run:
                        continue

                    row_db = db.execute(
                        select(ProjectProgress).where(
                            ProjectProgress.project_id == project.project_id,
                            ProjectProgress.task_id == task.task_id,
                        )
                    ).scalar_one_or_none()
                    if row_db is None:
                        row_db = ProjectProgress(
                            project_id=project.project_id,
                            task_id=task.task_id,
                        )
                        db.add(row_db)

                    row_db.status = status
                    if assigned_i is not None:
                        row_db.assigned_to = _cell_str(row[assigned_i])
                    if plan_s_i is not None:
                        row_db.planned_start = _cell_str(row[plan_s_i])
                    if plan_e_i is not None:
                        row_db.planned_end = _cell_str(row[plan_e_i])
                    if start_i is not None:
                        row_db.started_at = _cell_str(row[start_i])
                    if done_i is not None:
                        row_db.completed_at = _cell_str(row[done_i])
                    if vendor_i is not None:
                        row_db.vendor = _cell_str(row[vendor_i])
                    if blocker_i is not None:
                        row_db.blocker_note = (
                            _cell_str(row[blocker_i]) if status == "卡点" else None
                        )
                    if notes_i is not None:
                        row_db.notes = _cell_str(row[notes_i])

                    touched_project_ids.add(project.project_id)

    if pitfall_ws is not None:
        rows = list(pitfall_ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = _header_map(rows[0])
            stage_i = _col(headers, "stage_ref", "关联阶段", "阶段")
            task_i = _col(headers, "task_ref", "关联任务", "任务编号")
            wrong_i = _col(headers, "wrong_action", "错误做法")
            right_i = _col(headers, "right_action", "合规做法", "正确做法")
            if stage_i is None or task_i is None or wrong_i is None or right_i is None:
                summary.errors.append(
                    f"「{SHEET_PITFALLS}」缺少 stage_ref / task_ref / wrong_action / right_action 列"
                )
            else:
                impact_i = _col(headers, "impact_level", "影响等级")
                std_i = _col(headers, "standard_ref", "依据", "依据/规范")
                trig_i = _col(headers, "trigger_condition", "触发条件")
                rem_i = _col(headers, "remediation", "补救建议")
                notes_i = _col(headers, "notes", "备注")
                pit_id_i = _col(headers, "pitfall_id", "编号")

                stages_by_name = {
                    s.stage_name: s for s in db.execute(select(StageMap)).scalars().all()
                }
                tasks_by_code: dict[str, TaskDetail] = {
                    t.task_code.strip(): t
                    for t in db.execute(
                        select(TaskDetail).where(TaskDetail.is_active == 1)
                    ).scalars().all()
                    if t.task_code
                }

                for ridx, row in enumerate(rows[1:], start=2):
                    if not row:
                        continue
                    stage_ref = _cell_str(row[stage_i]) if stage_i < len(row) else None
                    task_ref = _cell_str(row[task_i]) if task_i < len(row) else None
                    wrong = _cell_str(row[wrong_i]) if wrong_i < len(row) else None
                    right = _cell_str(row[right_i]) if right_i < len(row) else None
                    if not stage_ref or not task_ref or not wrong or not right:
                        summary.pitfalls_skipped += 1
                        summary.warnings.append(
                            f"避坑第 {ridx} 行缺少必填字段（阶段/任务/错误做法/合规做法）"
                        )
                        continue
                    stage = stages_by_name.get(stage_ref)
                    if stage is None:
                        summary.pitfalls_skipped += 1
                        summary.warnings.append(
                            f"避坑第 {ridx} 行阶段不存在: {stage_ref}"
                        )
                        continue
                    task = tasks_by_code.get(task_ref)
                    if task is None:
                        summary.pitfalls_skipped += 1
                        summary.warnings.append(
                            f"避坑第 {ridx} 行任务不存在或已停用: {task_ref}"
                        )
                        continue
                    if task.stage_id != stage.stage_id:
                        summary.pitfalls_skipped += 1
                        summary.warnings.append(
                            f"避坑第 {ridx} 行任务 {task_ref} 与阶段 {stage_ref} 不一致"
                        )
                        continue
                    if _task_level(task_ref) not in (2, 3):
                        summary.pitfalls_skipped += 1
                        summary.warnings.append(
                            f"避坑第 {ridx} 行任务 {task_ref} 非二级或三级任务"
                        )
                        continue

                    impact = _cell_str(row[impact_i]) if impact_i is not None and impact_i < len(row) else "中"
                    standard = _cell_str(row[std_i]) if std_i is not None and std_i < len(row) else None
                    trigger = _cell_str(row[trig_i]) if trig_i is not None and trig_i < len(row) else None
                    remediation = _cell_str(row[rem_i]) if rem_i is not None and rem_i < len(row) else None
                    notes = _cell_str(row[notes_i]) if notes_i is not None and notes_i < len(row) else None

                    pit_id_raw = row[pit_id_i] if pit_id_i is not None and pit_id_i < len(row) else None
                    existing = None
                    if pit_id_raw is not None and str(pit_id_raw).strip():
                        try:
                            pid_int = int(str(pit_id_raw).strip())
                            existing = db.get(PitfallGuide, pid_int)
                        except (TypeError, ValueError):
                            pass
                    if existing is None:
                        summary.pitfalls_created += 1
                        if dry_run:
                            continue
                        p = PitfallGuide(
                            stage_ref=stage_ref,
                            task_ref=task_ref,
                            wrong_action=wrong,
                            right_action=right,
                            standard_ref=standard,
                            impact_level=impact or "中",
                            error_index=impact or "中",
                            trigger_condition=trigger,
                            remediation=remediation,
                            notes=notes,
                            source="导入",
                            verified=0,
                        )
                        db.add(p)
                    else:
                        summary.pitfalls_updated += 1
                        if dry_run:
                            continue
                        existing.stage_ref = stage_ref
                        existing.task_ref = task_ref
                        existing.wrong_action = wrong
                        existing.right_action = right
                        if standard is not None:
                            existing.standard_ref = standard
                        if impact:
                            existing.impact_level = impact
                            existing.error_index = impact
                        if trigger is not None:
                            existing.trigger_condition = trigger
                        if remediation is not None:
                            existing.remediation = remediation
                        if notes is not None:
                            existing.notes = notes

    if summary.errors:
        return summary

    if not dry_run:
        db.flush()
        for pid in touched_project_ids:
            project = db.get(ProjectProfile, pid)
            if project:
                recalculate_progress_percent(db, project)
                sync_project_status(db, project)

        log_action(
            db,
            actor,
            "IMPORT",
            "data",
            None,
            payload=summary.to_dict(),
            ip_address=ip_address,
            user_agent=user_agent,
        )
        db.commit()

    return summary
